import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
import re
import fitz  # PyMuPDF for PDF handling
import subprocess  # To open the Excel file after saving
from datetime import datetime
from openpyxl.styles import PatternFill

# Function to select folder using GUI
def select_folder():
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    folder_path = filedialog.askdirectory(title="Select Folder")
    return folder_path

# Function to scan folder and create an Excel log
def log_pdfs_to_excel(base_folder):
    # Create a new workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PDF Log"

    # Regular expression to find 10-digit part numbers in filenames
    part_number_regex = re.compile(r"\d{10}")

    # Dictionary to store part numbers and their locations (folders)
    part_numbers = {}
    latests = {}

    # List to store folder paths (columns) for Excel header
    folder_columns = []

    # Traverse folders and collect PDFs
    for root, dirs, files in os.walk(base_folder):
        folder_name = os.path.basename(root)
        pdfs = [f for f in files if f.lower().endswith(".pdf")]
        
        # Skip the base folder itself
        if root == base_folder:
            continue
        
        # Only create a column for folders that contain PDFs
        if pdfs:
            relative_folder_path = os.path.relpath(root, base_folder)
            folder_columns.append(relative_folder_path.replace(os.sep, '/'))  # Use '/' as separator

        # Scan each PDF file for part numbers
        for pdf_name in pdfs:
            match = part_number_regex.search(pdf_name)
            if match:
                part_number = match.group(0)  # Extract 10-digit part number
            else:
                match = re.compile(r"\d{9}").search(pdf_name)
                part_number = match.group(0) if match else pdf_name

            if part_number not in part_numbers:
                part_numbers[part_number] = {}
                latests[part_number] = []

            # Store the relative folder path (child folder)
            relative_path = os.path.relpath(root, base_folder).replace(os.sep, '/')
            pdf_path = os.path.join(root, pdf_name)

            timestamp = os.path.getmtime(pdf_path)
            last_modified_date = datetime.fromtimestamp(timestamp)
            if not latests[part_number]:
                latests[part_number] = (last_modified_date, relative_path)

            # Get revision number
            rev_no = None
            if "obsolete" in pdf_name.lower():
                rev_no = "obsolete"
            else:
                text = extract_text(pdf_path)
                for entry in text:
                    if entry["text"].strip() != "":
                        if "obsolete" in entry["text"].lower():
                            rev_no = "obsolete"
                            break
                if not rev_no:
                    rev_no = get_pdf_text(pdf_path, (1120, 772, 1163, 802))
                    rev_match = re.findall(r'[A-Z]?\.\d{2}', rev_no, re.IGNORECASE)
                    rev_no = rev_match[-1] if rev_match else ""

            # If revision number already exists for this part number and folder, append the new revision
            if not rev_no.strip():
                rev_match2 = re.findall(r'[A-Z]?\.\d{2}\.', pdf_name, re.IGNORECASE)
                rev_no = rev_match2[0][:-1] if rev_match2 else "no rev #"
            existing_revs = part_numbers[part_number].get(relative_path, [])
            if rev_no not in existing_revs:
                existing_revs.append(rev_no)
            part_numbers[part_number][relative_path] = existing_revs
            if latests[part_number]:
                existing_time, _ = latests[part_number]
                if last_modified_date > existing_time:
                    latests[part_number] = (last_modified_date, relative_path)

    # Write header to Excel
    sheet.cell(row=1, column=1, value="Part Number")  # Part number header
    for col, folder in enumerate(folder_columns, start=2):
        sheet.cell(row=1, column=col, value=folder)  # Folder/child folder headers

    # Write data to Excel
    row = 2  # Start at row 2 to leave space for header
    for part_number, locations in part_numbers.items():
        sheet.cell(row=row, column=1, value=part_number)  # Part number in the first column
        # Write the revision numbers for each folder
        for col, folder in enumerate(folder_columns, start=2):
            revs = locations.get(folder, [])
            cell = sheet.cell(row=row, column=col, value=", ".join(revs))  # Concatenate rev_no with commas
            # Check condition for highlighting
            if latests[part_number] and latests[part_number][1] == folder:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        row += 1

    # Save the Excel file and handle errors if file is open
    output_file = os.path.join(base_folder, "PDF_Log.xlsx")
    try:
        workbook.save(output_file)
        print(f"PDF log saved to: {output_file}")
        open_excel_file(output_file)  # Open the Excel file after saving
    except PermissionError:
        # If the file is open, show an error message and offer a retry button
        messagebox.showerror("File Open", f"The file '{output_file}' is currently open. Please close it and try again.")
        retry_prompt(output_file)

def retry_prompt(output_file):
    """Prompt the user to retry saving the file if it is open"""
    retry_response = messagebox.askretrycancel("Retry", f"Do you want to try saving the file '{output_file}' again?")
    if retry_response:
        log_pdfs_to_excel(os.path.dirname(output_file))  # Retry the save process

def open_excel_file(output_file):
    """Open the saved Excel file"""
    try:
        # For Windows, use subprocess to open the file with its default application
        subprocess.run(['start', '', output_file], shell=True, check=True)
    except Exception as e:
        print(f"Failed to open file: {e}")

def get_pdf_text(pdf_path, coordinates):
    x0, y0, x1, y1 = coordinates
    with fitz.open(pdf_path) as pdf:
        page = pdf[0]
        text = page.get_text("text", clip=fitz.Rect(x0, y0, x1, y1))
    return text

def extract_text(pdf_path):
    # Open the PDF file
    with fitz.open(pdf_path) as pdf:
        page_text_data = []  # Store data for current page
        # Iterate through each page in the PDF
        for page_num in range(len(pdf)):
            page = pdf[page_num]
            
            # Cycle through annotations to extract markup texts
            if page.annots():
                for annot in page.annots():
                    # Get annotation coordinates
                    rect = annot.rect
                    annot_x0, annot_y0, annot_x1, annot_y1 = rect.x0, rect.y0, rect.x1, rect.y1

                    # Extract text from the annotation and store it with page and coordinates
                    annot_text = annot.info.get("content", "")
                    page_text_data.append({
                        "page": page_num + 1,
                        "text": annot_text,
                        "coordinates": (annot_x0, annot_y0, annot_x1, annot_y1)
                    })

    return page_text_data

folder = select_folder()
if folder:
    log_pdfs_to_excel(folder)
else:
    print("No folder selected.")
